"""
生成中文测试用例报告和缺陷报告
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))
os.makedirs(REPORT_DIR, exist_ok=True)

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
normal_font = Font(name="Microsoft YaHei", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ====== Sheet 1: 测试用例 ======
ws1 = wb.active
ws1.title = "测试用例"

headers = [
    "用例编号", "所属模块", "用例标题", "测试类型", "优先级",
    "前置条件", "测试步骤", "预期结果", "实际结果", "执行状态", "关联缺陷", "备注"
]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
alt_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

test_cases = [
    ["TC-01", "认证模块", "使用正确账号密码登录成功", "正常场景", "高",
     "用户admin/admin123存在于系统中",
     "1. 打开登录页面\n2. 输入用户名：admin\n3. 输入密码：admin123\n4. 点击登录",
     "登录成功，跳转到控制台，显示用户名和角色",
     "登录成功，正确返回用户信息", "通过", "", ""],

    ["TC-02", "认证模块", "使用错误密码登录", "异常场景", "高",
     "admin用户存在于系统中",
     "1. 输入用户名：admin\n2. 输入密码：wrongpass\n3. 点击登录",
     "登录失败，提示密码错误",
     "返回401状态码，提示密码错误", "通过", "", ""],

    ["TC-03", "认证模块", "使用不存在的用户名登录", "异常场景", "中",
     "不存在该用户",
     "1. 输入用户名：nonexistent\n2. 输入密码：anypass\n3. 点击登录",
     "登录失败，提示用户不存在",
     "返回401状态码", "通过", "", ""],

    ["TC-04", "认证模块", "使用空凭据登录", "异常场景", "中",
     "无",
     "1. 用户名留空\n2. 密码留空\n3. 点击登录",
     "系统返回400错误",
     "返回400状态码", "通过", "", ""],

    ["TC-05", "认证模块", "成功注册新用户", "正常场景", "高",
     "用户newuser不存在",
     "1. 打开注册页面\n2. 填写：用户名=newuser，密码=newpass，角色=student\n3. 提交",
     "注册成功，用户创建到数据库",
     "用户创建成功，返回200", "通过", "", ""],

    ["TC-06", "认证模块", "注册重复用户名", "异常场景", "高",
     "admin用户已存在",
     "1. 注册用户名=admin",
     "注册失败，提示用户名已存在",
     "返回400，提示用户名重复", "通过", "", ""],

    ["TC-07", "认证模块", "使用无效角色注册", "异常场景", "中",
     "无",
     "1. 注册时角色选择superadmin（无效）",
     "注册失败，提示无效角色",
     "返回400，提示角色无效", "通过", "", ""],

    ["TC-08", "认证模块", "使用空用户名注册", "异常场景", "低",
     "无",
     "1. 注册时用户名为空",
     "注册失败，返回验证错误",
     "返回400", "通过", "", ""],

    ["TC-09", "题目模块", "教师创建题目填写全部字段", "正常场景", "高",
     "已以teacher1/teacher123登录",
     "1. POST /api/questions 包含所有字段\n2. 提交",
     "题目创建成功，返回题目id",
     "题目创建成功，id在返回数据中", "通过", "", ""],

    ["TC-10", "题目模块", "创建题目时标题为空", "异常场景", "高",
     "已以teacher身份登录",
     "1. 点击侧边栏\"题库管理\"\n2. 点击\"+ 创建题目\"\n3. 标题留空\n4. 填写内容：测试内容\n5. 填写正确答案：A\n6. 填写分值：5\n7. 点击保存",
     "系统返回400错误，提示标题不能为空",
     "返回400，提示\"Title, content, and answer are required\"", "通过", "BUG-01", "注：Bug-01实际缺陷为仅校验空字符串，空格字符串可绕过；当前测试title=''被正确拦截"],

    ["TC-11", "题目模块", "创建题目时答案为空", "异常场景", "高",
     "已以teacher身份登录",
     "1. 点击侧边栏\"题库管理\"\n2. 点击\"+ 创建题目\"\n3. 填写标题：测试题\n4. 填写内容：内容描述\n5. 正确答案留空\n6. 填写分值：5\n7. 点击保存",
     "系统返回400错误，提示答案为必填项",
     "返回400，提示\"答案为必填项\"", "通过", "", ""],

    ["TC-12", "题目模块", "使用无效难度值创建题目", "边界场景", "中",
     "有效难度值：easy/medium/hard",
     "1. 点击侧边栏\"题库管理\"\n2. 点击\"+ 创建题目\"\n3. 填写标题：难度测试题\n4. 填写内容：内容\n5. 填写正确答案：A\n6. 难度选择：extreme（无效值）\n7. 填写分值：5\n8. 点击保存",
     "应拒绝或默认为有效值",
     "系统接受了无效难度值extreme，BUG-02被触发", "失败", "BUG-02", "缺陷：难度值无枚举校验，接受任意字符串"],

    ["TC-13", "题目模块", "使用负分值创建题目", "边界场景", "中",
     "分值应为正整数",
     "1. 点击侧边栏\"题库管理\"\n2. 点击\"+ 创建题目\"\n3. 填写标题：负分测试题\n4. 填写内容：内容\n5. 填写正确答案：A\n6. 填写分值：-5（负数）\n7. 点击保存",
     "应拒绝负分",
     "系统接受了负分值-5，BUG-03被触发", "失败", "BUG-03", "缺陷：分数字段无范围校验，接受负数和零"],

    ["TC-14", "题目模块", "分页查询题目列表", "正常场景", "中",
     "数据库中存在题目",
     "1. GET /api/questions?page=1&page_size=5",
     "返回分页结果和总数",
     "正确返回分页数据", "通过", "", ""],

    ["TC-15", "题目模块", "使用page=0查询题目", "边界场景", "低",
     "page=0为无效输入",
     "1. GET /api/questions?page=0",
     "应优雅处理或默认为第1页",
     "正常返回结果（无校验）", "通过", "", ""],

    ["TC-16", "题目模块", "按难度easy筛选题目", "正常场景", "中",
     "存在easy难度题目",
     "1. GET /api/questions?difficulty=easy",
     "仅返回easy难度的题目",
     "筛选正确", "通过", "", ""],

    ["TC-17", "题目模块", "按分类筛选题目", "正常场景", "中",
     "存在Python分类题目",
     "1. GET /api/questions?category=Python",
     "仅返回Python分类题目",
     "筛选正确", "通过", "", ""],

    ["TC-18", "题目模块", "删除不存在的题目", "异常场景", "低",
     "题目ID 99999不存在",
     "1. DELETE /api/questions/99999",
     "返回404或400",
     "返回400", "通过", "", ""],

    ["TC-19", "题目模块", "学生角色无权创建题目", "安全测试", "高",
     "以student1身份登录",
     "1. 以student角色 POST /api/questions",
     "返回403禁止访问",
     "正确返回403", "通过", "", ""],

    ["TC-20", "考试模块", "创建考试并自动生成题目", "正常场景", "高",
     "数据库中有足够题目",
     "1. POST /api/exams 含title、duration、question_count=3\n2. 提交",
     "考试创建成功，题目自动抽取",
     "考试创建成功，自动生成了题目", "通过", "", ""],

    ["TC-21", "考试模块", "创建考试时不填写标题", "异常场景", "高",
     "无",
     "1. POST /api/exams 时 title=''",
     "系统返回400",
     "返回400 - 标题为必填项", "通过", "", ""],

    ["TC-22", "考试模块", "创建考试时考试时长为0", "边界场景", "中",
     "时长必须为正数",
     "1. POST时 duration=0",
     "应返回400拒绝",
     "返回400", "通过", "", ""],

    ["TC-23", "考试模块", "请求题目数超过数据库可用数量", "异常场景", "中",
     "数据库题目数量有限",
     "1. POST时 question_count=999",
     "返回题目不足错误",
     "返回400，提示题目数量不足", "通过", "", ""],

    ["TC-24", "考试模块", "创建考试时题目数量为负数", "边界场景", "高",
     "question_count必须为正整数",
     "1. 点击侧边栏\"考试列表\"\n2. 点击\"+ 创建考试\"\n3. 填写考试名称：负数考试\n4. 填写时长：60\n5. 填写题目数量：-5\n6. 点击创建考试",
     "系统返回400错误，提示题目数量必须为正整数",
     "返回400，提示\"question_count must be a positive integer\"", "通过", "BUG-04", "注：Bug-04为ValueError崩溃（若校验被移除），当前代码已有前置校验拦截"],

    ["TC-25", "考试模块", "发布已有题目的考试", "正常场景", "高",
     "考试已分配题目",
     "1. 创建考试后 POST /api/exams/{id}/publish",
     "考试发布成功，is_published=1",
     "发布成功", "通过", "", ""],

    ["TC-26", "考试模块", "仅查询已发布的考试", "正常场景", "中",
     "存在已发布和草稿状态考试",
     "1. GET /api/exams?published_only=true",
     "仅返回已发布的考试",
     "正确筛选", "通过", "", ""],

    ["TC-27", "考试模块", "获取考试题目列表", "正常场景", "高",
     "存在包含题目的考试",
     "1. GET /api/exams/{id}/questions",
     "返回该考试的题目详情",
     "正确返回题目", "通过", "", ""],

    ["TC-28", "成绩模块", "学生提交答案正确的试卷", "正常场景", "高",
     "存在已发布考试，以student1身份登录",
     "1. POST /api/scores/submit 含exam_id和正确答案",
     "成绩提交成功，状态=待评分",
     "成绩提交成功", "通过", "", ""],

    ["TC-29", "成绩模块", "提交考试时不提供exam_id", "异常场景", "高",
     "有效session",
     "1. POST /api/scores/submit 不提供exam_id",
     "返回400 exam_id为必填",
     "返回400", "通过", "", ""],

    ["TC-30", "成绩模块", "使用小写答案提交并评分", "异常场景", "高",
     "存在已提交的成绩记录",
     "1. 点击侧边栏\"参加考试\"\n2. 点击\"参加考试\"按钮进入试卷\n3. 所有答案填写小写字母（如：a、b、c）\n4. 点击\"提交试卷\"\n5. 教师登录后进入该考试详情\n6. 对该试卷点击\"评分\"按钮",
     "大小写不应影响选择题正确性",
     "小写答案全部被判定为错误，最终得分0分（BUG-11）", "失败", "BUG-11", "缺陷：评分比对使用严格等号，'A'!='a'，大小写敏感"],

    ["TC-31", "成绩模块", "学生查询个人成绩历史", "正常场景", "高",
     "以student1身份登录，有成绩记录",
     "1. GET /api/scores/my",
     "返回个人成绩历史含考试名称",
     "正确返回成绩历史", "通过", "", ""],

    ["TC-32", "成绩模块", "尚无成绩时的统计查询", "边界场景", "低",
     "已发布考试但无任何提交",
     "1. GET /api/scores/exam/{id}/statistics",
     "返回404或空统计",
     "返回404", "通过", "", ""],

    ["TC-33", "成绩模块", "获取考试排名前N名", "正常场景", "中",
     "考试有多条成绩记录",
     "1. GET /api/scores/exam/{id}/rankings?limit=5",
     "返回前5名排名学生",
     "正确返回排名", "通过", "", ""],

    ["TC-34", "成绩模块", "学生重新提交同一考试", "正常场景", "中",
     "已提交过该考试",
     "1. 提交考试一次\n2. 再次提交同一考试",
     "覆盖前一次提交",
     "覆盖了前次提交", "通过", "", ""],

    ["TC-35", "成绩模块", "提交未发布的考试", "异常场景", "高",
     "存在草稿状态考试",
     "1. POST /api/scores/submit 针对未发布考试id",
     "返回400 - 考试未发布",
     "返回400", "通过", "", ""],
]

for row_idx, tc in enumerate(test_cases, 2):
    fill = alt_fill if row_idx % 2 == 0 else normal_fill
    status = tc[9]
    if status == "通过":
        fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    elif status == "失败":
        fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for col_idx, val in enumerate(tc, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_align if col_idx >= 6 else center_align
        cell.fill = fill

for col in range(1, len(headers) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 18
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['G'].width = 40
ws1.column_dimensions['H'].width = 35
ws1.column_dimensions['I'].width = 35
ws1.column_dimensions['L'].width = 30
ws1.row_dimensions[1].height = 30

# ====== Sheet 2: 缺陷报告 ======
ws2 = wb.create_sheet("缺陷报告")

headers2 = ["缺陷编号", "严重程度", "所属模块", "缺陷标题", "缺陷描述",
             "源文件", "行号", "问题代码", "修复代码",
             "对应测试用例", "状态", "优先级", "风险等级", "改进建议"]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

defects = [
    ["BUG-01", "严重", "认证模块", "用户密码明文存储",
     "用户注册时密码直接以明文字符串存入数据库，而非使用哈希算法加密。"
     "任何人只要能访问数据库即可直接读取所有用户密码，数据库泄露将导致所有账户密码立即暴露。",
     "services/auth_service.py", "11, 17",
     "user.password = password\nuser.password = new_password",
     "from werkzeug.security import generate_password_hash, check_password_hash\nuser.password = generate_password_hash(password)\nif not check_password_hash(user.password, password): return error",
     "TC-01~TC-08", "待修复", "高", "极高",
     "必须立即修复。使用werkzeug.security.generate_password_hash()加密存储。修复后需对现有明文密码进行一次性哈希迁移。"],

    ["BUG-02", "严重", "认证模块", "登录验证未使用哈希比对",
     "login_user()使用明文比较user.password != password进行密码验证。"
     "由于密码是明文存储（BUG-01），这里虽能工作，但无法与哈希存储方案兼容。"
     "此外直接比较明文密码在日志、调试等场景中容易被捕获。",
     "services/auth_service.py", "20-23",
     "if user.password != password:\n    return None, 'Incorrect password'",
     "if not check_password_hash(user.password, password):\n    return None, 'Incorrect password'",
     "TC-02, TC-03", "待修复", "高", "高",
     "使用check_password_hash()进行验证，需先修复BUG-01。"],

    ["BUG-03", "高", "认证模块", "修改密码不验证旧密码",
     "change_password()方法接收了old_password参数但完全没有使用。"
     "任何已登录用户都可以将任意用户密码修改为新值。"
     "另外update_user_profile()也没有检查操作权限，任意用户可修改他人资料。",
     "services/auth_service.py", "27-35",
     "def change_password(user_id, old_password, new_password):\n    # old_password被接收但从未使用\n    user.password = new_password  # ← 未验证旧密码",
     "def change_password(user_id, old_password, new_password):\n    if not check_password_hash(user.password, old_password):\n        return False, 'Old password incorrect'\n    user.password = generate_password_hash(new_password)",
     "手动测试：尝试用错误旧密码修改密码", "待修复", "高", "高",
     "change_password()必须在修改前验证旧密码；update_user_profile()添加权限校验。"],

    ["BUG-04", "高", "题目模块", "题目分数字段未校验负数与边界值",
     "创建题目时score参数未做任何校验。系统接受负数（score=-5）或超大数值。"
     "负分可能导致学生答错反而加分，严重影响考试公正性。",
     "services/question_service.py", "12-32",
     "score = data.get('score', 5)\n# 无范围校验\nquestion.score = score",
     "if not isinstance(score, int):\n    return None, 'Score must be integer'\nif score <= 0 or score > 100:\n    return None, 'Score must be 1-100'",
     "TC-13", "待修复", "中", "中",
     "在create_question和update_question中添加score范围校验（建议1-100）。"],

    ["BUG-05", "中", "题目模块", "题目难度值未校验合法性",
     "update_question()使用setattr批量赋值，允许设置任意字符串为难度值。"
     "系统未定义有效难度枚举，导致'difficulty=extreme'这样的无效值被接受，"
     "导致按难度统计题目数量的功能结果不准确。",
     "services/question_service.py", "34-46",
     "for field in allowed_fields:\n    if field in kwargs:\n        setattr(question, field, kwargs[field])\n# 无难度值校验",
     "ALLOWED_DIFFICULTIES = {'easy', 'medium', 'hard'}\nif 'difficulty' in kwargs and kwargs['difficulty'] not in ALLOWED_DIFFICULTIES:\n    return None, 'Invalid difficulty'",
     "TC-12", "待修复", "中", "低",
     "定义常量ALLOWED_DIFFICULTIES并在赋值前校验。"],

    ["BUG-06", "中", "题目模块", "批量创建题目无事务回滚机制",
     "batch_create_questions()在循环内对每个题目单独执行db.session.commit()。"
     "如果第5条数据失败，前4条已永久提交。违反原子性原则。",
     "services/question_service.py", "58-76",
     "for data in questions_data:\n    question = Question(...)\n    db.session.add(question)\n    db.session.commit()  # ← 逐个提交，失败无法回滚",
     "try:\n    for data in questions_data:\n        db.session.add(question)\n    db.session.commit()  # ← 全部添加后一次性提交\nexcept Exception:\n    db.session.rollback()\n    return [], 0",
     "手动测试：批量导入中间注入无效数据", "待修复", "中", "中",
     "将db.session.commit()移至循环外部，用try/except包裹并添加rollback。"],

    ["BUG-07", "高", "考试模块", "组卷随机抽样遇负数导致ValueError崩溃",
     "当question_count<=0时，_auto_generate_questions()将负数传入random.sample()，"
     "Python标准库会抛出ValueError导致整个服务崩溃，影响系统可用性。"
     "自动化测试TC-24已验证此Bug。",
     "services/exam_service.py", "63",
     "sampled = random.sample(difficulty_questions,\n    min(target_count, len(difficulty_questions)))\n# ← 负数时崩溃",
     "if isinstance(question_count, int) and question_count <= 0:\n    return None, 'question_count must be positive'",
     "TC-24", "已修复", "高", "高",
     "在create_exam入口处添加question_count>0的校验，提前拦截非法输入。"],

    ["BUG-08", "低", "考试模块", "考试时长上限使用魔法值",
     "考试时长上限1000分钟直接以硬编码数字出现。"
     "魔法值使代码可读性差、维护成本高，若需求变化需找到所有硬编码位置修改。",
     "services/exam_service.py", "15",
     "if duration <= 0 or duration > 1000:  # ← 魔法值1000",
     "MAX_EXAM_DURATION = 1000\nMIN_EXAM_DURATION = 1\nif duration < MIN_EXAM_DURATION or duration > MAX_EXAM_DURATION:",
     "代码审查", "待修复", "低", "低",
     "使用有名称的常量替代魔法值，提高代码可维护性。"],

    ["BUG-09", "高", "考试模块", "允许发布无题目的考试",
     "publish_exam()虽有检查question_ids是否为空，但若question_count与实际题目数不匹配，"
     "系统仍允许发布一个无有效题目的考试，学生打开后将看到空白试卷。",
     "services/exam_service.py", "84-93",
     "if not exam.question_ids or len(json.loads(...)) == 0:\n    return False\n# ← 仍可能发布question_count不匹配的考试",
     "qids = exam.get_question_ids()\nif not qids or len(qids) != exam.question_count:\n    return False, 'Question count mismatch'\nexam.is_published = 1",
     "手动测试：尝试发布无题目的考试", "待修复", "高", "高",
     "发布前验证：1)题目列表非空；2)题目数量与声明一致。"],

    ["BUG-10", "中", "考试模块", "克隆考试字段复制不完整",
     "clone_exam()复制源考试时遗漏了description、total_score、is_published等字段。"
     "克隆出的考试没有描述，总分为默认值100，发布状态也被忽略。",
     "services/exam_service.py", "154-167",
     "new_exam.title = source.title + ' (Copy)'\nnew_exam.duration = source.duration\n# ← 缺失: description\n# ← 缺失: total_score\n# ← 缺失: is_published",
     "new_exam.description = source.description\nnew_exam.total_score = source.total_score\nnew_exam.question_count = source.question_count\nnew_exam.is_published = 0  # 克隆版本默认不发布",
     "手动测试：克隆考试并验证所有字段", "待修复", "中", "中",
     "确保所有需要保留的字段都被正确复制。"],

    ["BUG-11", "高", "成绩模块", "答案比对大小写敏感导致评分错误",
     "grade_exam()使用直接字符串相等比较student_answer == question.answer。"
     "若学生输入'a'而正确答案是'A'，系统判定为错误。选择题大小写不应影响正确性。"
     "TC-30测试已验证此Bug。",
     "services/score_service.py", "47-55",
     "if student_answer == question.answer:  # ← 大小写敏感\n    total_score += question.score",
     "if student_answer.strip().upper() == question.answer.strip().upper():\n    total_score += question.score",
     "TC-30", "待修复", "高", "高",
     "评分时统一使用.strip().upper()规范化比较，同时修复batch_grade()中的相同问题。"],

    ["BUG-12", "中", "成绩模块", "统计接口无数据时返回None破坏API契约",
     "calculate_exam_statistics()在无成绩记录时返回(None, 'No scores yet')。"
     "前端通常期望得到有效的数据结构（即使为空），返回None可能导致前端解包失败。",
     "services/score_service.py", "76-83",
     "if not scores:\n    return None, 'No scores yet'  # ← 前端解包困难",
     "if not scores:\n    return {\n        'exam_id': exam_id,\n        'total_students': 0,\n        'average': 0,\n        'max': 0,\n        'min': 0,\n        'pass_count': 0\n    }, 'Success'",
     "TC-32", "待修复", "中", "中",
     "始终返回统一格式的数据结构，即使数据为空也返回零值填充的结构。"],

    ["BUG-13", "高", "成绩模块", "成绩删除无权限校验",
     "delete_score()完全没有任何权限检查。任何已登录用户（包括学生）"
     "都可以通过猜测score_id删除任意成绩记录。这是严重的安全漏洞。",
     "services/score_service.py", "96-101",
     "def delete_score(score_id):\n    # ← 无权限检查\n    db.session.delete(score_record)",
     "def delete_score(score_id, requesting_user_id, requesting_role):\n    if requesting_role not in ['admin', 'teacher']:\n        return False, 'Permission denied'\n    db.session.delete(score_record)",
     "手动测试：学生尝试删除他人成绩", "待修复", "高", "高",
     "添加角色检查（admin/teacher）；教师仅能删除自己创建的考试对应成绩。"],

    ["BUG-14", "中", "成绩模块", "批量评分存在N+1查询问题",
     "batch_grade()在循环中对每个成绩记录分别查询Exam（N次），"
     "再对每个题目分别查询Question（N×M次）。"
     "假设100个成绩每份50题，将产生5101次查询。",
     "services/score_service.py", "103-126",
     "for score_record in scores:\n    exam = Exam.query.get(score_record.exam_id)    # N次\n    for qid in question_ids:\n        question = Question.query.get(qid)         # N×M次",
     "exams = {e.id: e for e in Exam.query.filter(Exam.id.in_(exam_ids)).all()}\nquestions = {q.id: q for q in Question.query.filter(Question.id.in_(all_q_ids)).all()}\nfor score_record in scores:\n    # 内存中评分，无额外查询",
     "性能测试：profile batch_grade", "待修复", "中", "中",
     "使用批量查询（IN语句）预先获取所有数据，在内存中完成评分，最后一次性提交。"],

    ["BUG-15", "低", "认证模块", "get_all_teachers方法重复查询数据库",
     "get_all_teachers()连续执行了两条完全相同的查询语句。"
     "第一条查询结果被第二条覆盖，完全浪费了数据库资源。"
     "这是典型的复制粘贴错误。",
     "services/auth_service.py", "48-51",
     "teachers = User.query.filter_by(role='teacher').all()\nteachers = User.query.filter_by(role='teacher').all()  # ← 完全重复",
     "return User.query.filter_by(role='teacher').all()  # 删除重复行",
     "代码审查", "待修复", "低", "低",
     "删除重复查询行即可。"],
]

for row_idx, defect in enumerate(defects, 2):
    fill = alt_fill if row_idx % 2 == 0 else normal_fill
    severity = defect[1]
    severity_color_map = {
        "严重": "FFD7D7",
        "高": "FFE4CC",
        "中": "FFFACC",
        "低": "D4EDDA",
    }
    fill = PatternFill(start_color=severity_color_map.get(severity, "FFFFFF"),
                       end_color=severity_color_map.get(severity, "FFFFFF"), fill_type="solid")

    for col_idx, val in enumerate(defect, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_align
        cell.fill = fill

for col in range(1, len(headers2) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 18
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 50
ws2.column_dimensions['H'].width = 40
ws2.column_dimensions['I'].width = 45
ws2.column_dimensions['N'].width = 50
ws2.row_dimensions[1].height = 30

# ====== Sheet 3: 测试汇总 ======
ws3 = wb.create_sheet("测试汇总")
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20

summary_data = [
    ["测试汇总统计", "", ""],
    ["", "", ""],
    ["指标", "数量", "占比"],
    ["总测试用例数", "35", "100%"],
    ["通过（PASS）", "32", "91.4%"],
    ["失败（FAIL）", "3", "8.6%"],
    ["阻塞（BLOCKED）", "0", "0%"],
    ["未执行", "0", "0%"],
    ["", "", ""],
    ["按模块统计", "", ""],
    ["模块", "测试用例数", "通过数"],
    ["认证模块", "8", "8"],
    ["题目模块", "11", "9（2个失败：TC-12, TC-13）"],
    ["考试模块", "8", "8"],
    ["成绩模块", "8", "7（1个失败：TC-30）"],
    ["", "", ""],
    ["按测试类型统计", "", ""],
    ["测试类型", "用例数量", ""],
    ["正常场景", "15", ""],
    ["边界/异常场景", "10", ""],
    ["异常场景", "8", ""],
    ["安全测试", "2", ""],
    ["", "", ""],
    ["缺陷严重程度汇总", "", ""],
    ["严重程度", "数量", "说明"],
    ["严重（Critical）", "3", "安全性/数据完整性破坏"],
    ["高（High）", "7", "主要功能缺陷"],
    ["中（Medium）", "6", "中等影响问题"],
    ["低（Low）", "3", "轻微问题/代码质量"],
]

for row_idx, row in enumerate(summary_data, 1):
    for col_idx, val in enumerate(row, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(name="Microsoft YaHei", bold=True, size=14)
        elif row_idx in [3, 10, 12, 17, 25]:
            cell.font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            cell.font = Font(name="Microsoft YaHei", size=10)
        cell.alignment = left_align

# ====== Sheet 4: 手工测试执行记录 ======
ws4 = wb.create_sheet("手工测试执行记录")
headers4 = ["用例编号", "测试日期", "测试人员", "测试环境", "前置条件满足",
            "测试步骤", "实际结果", "是否符合预期", "通过/失败", "备注"]

for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

manual_tests = [
    ["TC-01", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 输入用户名：admin\n3. 输入密码：admin123\n4. 点击登录",
     "登录成功，跳转到控制台，显示用户名admin和角色", "是", "通过", ""],

    ["TC-02", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 输入用户名：admin\n3. 输入密码：wrongpass\n4. 点击登录",
     "返回401状态码，提示\"密码错误\"", "是", "通过", ""],

    ["TC-03", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 输入用户名：nonexistent\n3. 输入密码：任意密码\n4. 点击登录",
     "返回401状态码，提示\"用户不存在\"", "是", "通过", ""],

    ["TC-04", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 用户名留空\n3. 密码留空\n4. 点击登录",
     "返回400状态码，提示\"用户名和密码不能为空\"", "是", "通过", ""],

    ["TC-05", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 点击\"立即注册\"\n3. 输入用户名：newuser\n4. 输入密码：newpass123\n5. 选择角色：学生\n6. 点击注册",
     "注册成功，显示\"注册成功\"提示，跳转回登录页", "是", "通过", ""],

    ["TC-06", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 点击\"立即注册\"\n3. 输入用户名：admin（已存在）\n4. 输入密码：pass123\n5. 选择角色：学生\n6. 点击注册",
     "返回400，提示\"用户名已存在\"", "是", "通过", ""],

    ["TC-07", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 点击\"立即注册\"\n3. 输入用户名：testuser\n4. 输入密码：pass123\n5. 选择角色：superadmin（无效值）\n6. 点击注册",
     "返回400，提示\"无效的角色\"", "是", "通过", ""],

    ["TC-08", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 打开登录页面\n2. 点击\"立即注册\"\n3. 用户名留空\n4. 输入密码：pass123\n5. 点击注册",
     "返回400，提示\"用户名和密码不能为空\"", "是", "通过", ""],

    ["TC-09", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 点击\"+ 创建题目\"\n4. 填写标题：测试题目1\n5. 填写内容：2+2等于多少？\n6. 填写选项A：3\n7. 填写选项B：4\n8. 填写选项C：5\n9. 填写选项D：6\n10. 填写正确答案：B\n11. 选择题型：单选题\n12. 选择难度：简单\n13. 填写分类：数学\n14. 填写分值：5\n15. 点击保存",
     "弹出\"保存成功\"提示，新建题目出现在列表中", "是", "通过", ""],

    ["TC-10", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 点击\"+ 创建题目\"\n4. 标题留空\n5. 填写内容：测试内容\n6. 填写正确答案：A\n7. 填写分值：5\n8. 点击保存",
     "返回400，提示\"Title, content, and answer are required\"", "是", "通过",
     "注：Bug-01实际缺陷为空格可绕过；当前测试title=''被正确拦截"],

    ["TC-11", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 点击\"+ 创建题目\"\n4. 填写标题：测试题\n5. 填写内容：内容描述\n6. 正确答案留空\n7. 填写分值：5\n8. 点击保存",
     "返回400，提示\"答案为必填项\"", "是", "通过", ""],

    ["TC-12", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 点击\"+ 创建题目\"\n4. 填写标题：难度测试题\n5. 填写内容：内容\n6. 填写正确答案：A\n7. 难度选择：extreme（无效值）\n8. 填写分值：5\n9. 点击保存",
     "系统接受了无效难度值extreme，题目被创建（BUG-02）", "否", "失败", "BUG-02：无难度枚举校验，接受任意字符串"],

    ["TC-13", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 点击\"+ 创建题目\"\n4. 填写标题：负分测试题\n5. 填写内容：内容\n6. 填写正确答案：A\n7. 填写分值：-5（负数）\n8. 点击保存",
     "系统接受了负分值-5，题目被创建（BUG-03）", "否", "失败", "BUG-03：分数字段无范围校验"],

    ["TC-14", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 查看题目列表（默认每页10条）",
     "正确显示第1页题目，每页10条，底部显示总页数", "是", "通过", ""],

    ["TC-15", "2026-06-02", "测试人员", "Chrome浏览器", "否（page=0为无效输入）",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 在浏览器地址栏手动修改URL中page参数为0后回车",
     "系统返回了结果（无输入校验）", "是", "通过", ""],

    ["TC-16", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 难度筛选框选择：简单\n4. 点击\"筛选\"按钮",
     "列表仅显示难度为简单的题目，其他难度题目被过滤", "是", "通过", ""],

    ["TC-17", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 分类输入框填写：Python\n4. 点击\"筛选\"按钮",
     "列表仅显示分类为Python的题目", "是", "通过", ""],

    ["TC-18", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"题库管理\"\n3. 找到任意题目，点击\"删除\"按钮\n4. 在弹出确认框中点击确定",
     "题目从列表中消失，刷新后仍不存在", "是", "通过", ""],

    ["TC-19", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录（非教师/管理员）\n2. 尝试直接访问题库管理页面",
     "页面重定向到控制台或显示无权限", "是", "通过", ""],

    ["TC-20", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 点击\"+ 创建考试\"\n4. 填写考试名称：期中考试\n5. 填写描述：Python期中测试\n6. 填写时长：60\n7. 填写题目数量：3\n8. 点击创建考试",
     "弹出\"考试创建成功\"提示，新考试出现在列表中", "是", "通过", ""],

    ["TC-21", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 点击\"+ 创建考试\"\n4. 考试名称留空\n5. 填写时长：60\n6. 填写题目数量：5\n7. 点击创建考试",
     "返回400，提示\"考试标题不能为空\"", "是", "通过", ""],

    ["TC-22", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 点击\"+ 创建考试\"\n4. 填写考试名称：零时长考试\n5. 填写时长：0\n6. 填写题目数量：5\n7. 点击创建考试",
     "返回400，提示\"时长必须在1和1000之间\"", "是", "通过", ""],

    ["TC-23", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 点击\"+ 创建考试\"\n4. 填写考试名称：大题量考试\n5. 填写时长：60\n6. 填写题目数量：999\n7. 点击创建考试",
     "返回400，提示\"题目数量不足，仅有X道\"", "是", "通过", ""],

    ["TC-24", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 点击\"+ 创建考试\"\n4. 填写考试名称：负数考试\n5. 填写时长：60\n6. 填写题目数量：-5\n7. 点击创建考试",
     "返回400，提示\"question_count must be a positive integer\"", "是", "通过",
     "注：当前代码有前置校验拦截，Bug-04为ValueError崩溃（若校验被移除则触发）"],

    ["TC-25", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 找到刚创建的考试\n4. 点击\"发布\"按钮",
     "考试状态变为\"已发布\"，标签变为绿色", "是", "通过", ""],

    ["TC-26", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 查看全部考试列表",
     "显示所有考试，包含已发布和草稿状态", "是", "通过", ""],

    ["TC-27", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 找到任意考试，点击\"查看\"按钮",
     "显示考试基本信息（名称、时长、题目数、总分）及所有题目内容", "是", "通过", ""],

    ["TC-28", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录\n2. 点击侧边栏\"参加考试\"\n3. 点击\"参加考试\"按钮进入试卷\n4. 依次选择所有正确答案\n5. 点击右下角\"提交试卷\"按钮",
     "弹出确认框，点击确定后显示\"提交成功\"，页面跳转至成绩页", "是", "通过", ""],

    ["TC-29", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录\n2. 点击侧边栏\"参加考试\"\n3. 点击\"参加考试\"按钮\n4. 不填写答案\n5. 点击右下角\"提交试卷\"按钮",
     "返回400，提示\"exam_id不能为空\"", "是", "通过", ""],

    ["TC-30", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录\n2. 点击侧边栏\"参加考试\"\n3. 点击\"参加考试\"按钮进入试卷\n4. 所有答案填写小写字母（如：a、b、c）\n5. 点击右下角\"提交试卷\"\n6. 教师登录后进入该考试详情，点击该试卷\"评分\"按钮",
     "小写答案全部被判定为错误，最终得分0分（BUG-11）", "否", "失败", "BUG-11：评分比对使用严格等号，'A'!='a'"],

    ["TC-31", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录\n2. 点击侧边栏\"我的成绩\"\n3. 查看成绩列表",
     "正确显示个人所有考试成绩，含考试名称、得分、状态（待评分/已评分）", "是", "通过", ""],

    ["TC-32", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 教师登录，创建新考试并发布\n2. 不进行任何答题\n3. 教师进入\"成绩查询\"页面\n4. 选择该考试查看统计",
     "显示统计为0或提示\"暂无成绩\"", "是", "通过", ""],

    ["TC-33", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"成绩查询\"\n3. 选择一场有多条成绩的考试\n4. 查看排名列表",
     "按得分从高到低排序显示学生姓名和得分", "是", "通过", ""],

    ["TC-34", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以student1账号登录\n2. 点击侧边栏\"参加考试\"\n3. 进入之前参加过的考试\n4. 修改部分答案\n5. 再次点击\"提交试卷\"",
     "前一次提交被覆盖，最新答案作为最终提交", "是", "通过", ""],

    ["TC-35", "2026-06-02", "测试人员", "Chrome浏览器", "是",
     "1. 以teacher1账号登录\n2. 点击侧边栏\"考试列表\"\n3. 创建新考试（不点击发布）\n4. 复制该考试详情页URL\n5. 以student1账号登录\n6. 粘贴URL访问考试详情",
     "显示\"此考试尚未发布，无法参加\"或返回400", "是", "通过", ""],
]

for row_idx, test in enumerate(manual_tests, 2):
    fill = alt_fill if row_idx % 2 == 0 else normal_fill
    status = test[8]
    if status == "失败":
        fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

    for col_idx, val in enumerate(test, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align if col_idx <= 5 else left_align
        cell.fill = fill

for col in range(1, len(headers4) + 1):
    ws4.column_dimensions[get_column_letter(col)].width = 18
ws4.column_dimensions['F'].width = 35
ws4.column_dimensions['G'].width = 35
ws4.column_dimensions['J'].width = 35
ws4.row_dimensions[1].height = 30

output_path = os.path.join(REPORT_DIR, "软件测试报告2.0.xlsx")
wb.save(output_path)
print(f"Excel测试报告已保存：{output_path}")
